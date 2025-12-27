"""
Excel Formula Extractor for Opportuniteitskost Calculator
Extracts formulas and structure from the FIRE calculator spreadsheet
"""

import sys
try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

def extract_sheet_data(ws, sheet_name):
    """Extract all data and formulas from a worksheet"""
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print('='*60)
    
    # Find the used range
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"Dimensions: {max_row} rows x {max_col} columns\n")
    
    # Extract all cells with values or formulas
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None or (hasattr(cell, 'value') and cell.value == 0):
                col_letter = cell.column_letter
                value = cell.value
                
                # Check if it's a formula
                if isinstance(value, str) and value.startswith('='):
                    row_data.append(f"{col_letter}{row}: FORMULA: {value}")
                elif value is not None:
                    row_data.append(f"{col_letter}{row}: {repr(value)}")
        
        if row_data:
            print(f"Row {row}:")
            for item in row_data:
                print(f"  {item}")

def main():
    excel_file = "../Opportuniteitskost - FIRE calculator.xlsx"
    
    print("Loading Excel file...")
    
    # Load with data_only=False to get formulas
    wb_formulas = load_workbook(excel_file, data_only=False)
    
    # Load with data_only=True to get calculated values
    wb_values = load_workbook(excel_file, data_only=True)
    
    print(f"Sheets found: {wb_formulas.sheetnames}")
    
    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        extract_sheet_data(ws, sheet_name)
    
    print("\n\n" + "="*60)
    print("CALCULATED VALUES (for reference)")
    print("="*60)
    
    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        print(f"\n--- {sheet_name} ---")
        for row in range(1, ws.max_row + 1):
            row_vals = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    row_vals.append(f"{cell.column_letter}{row}={cell.value}")
            if row_vals:
                print(f"  {row_vals}")

if __name__ == "__main__":
    main()

